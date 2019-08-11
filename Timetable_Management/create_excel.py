from flask import Flask
from flask import send_file
import openpyxl
from openpyxl import load_workbook,Workbook

def excel_batch(batch_info,shift_info,batch_data):
    """ Preparing excel sheet for batch wise timetable
        
        Args:
        1. batch_info - Batch inputted by user(FE,SE,TE,BE,etc)
        2. shift_info - SHift number inputted by user(1st shift OR 2nd shift)
        3. batch_data - Data from master timetable of particular batch_info and shift_info
    """
    Main=[]
    print(batch_info)
    print(shift_info)
    if batch_info == 'FE' and shift_info == '1':
        loc = './excel/FE1.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row,col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1

        xfile.save(loc)
        

    elif batch_info == 'FE' and shift_info == '2':
        loc = './excel/FE2.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row, col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1
        xfile.save(loc)
        

    elif batch_info == 'SE' and shift_info == '1':
        print('hey this is SE1')
        loc = './excel/SE1.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row, col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1
        xfile.save(loc)
        

    elif batch_info == 'SE' and shift_info == '2':
        loc = './excel/SE2.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row, col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1
        xfile.save(loc) 
        

    elif batch_info== 'TE' and shift_info == '1':
        loc = './excel/TE1.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row, col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1
        xfile.save(loc)
        

    elif batch_info == 'TE' and shift_info == '2':
        loc = './excel/TE2.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row, col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1
        xfile.save(loc)
        


    elif batch_info == 'BE' and shift_info == '1':
        loc = './excel/BE1.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row, col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1
        xfile.save(loc)
        

    elif batch_info == 'BE' and shift_info == '2':
        loc = './excel/BE2.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row, col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1
        xfile.save(loc)
        

    elif batch_info == 'ME' and shift_info == '1':
        loc = './excel/ME1.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row, col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1
        xfile.save(loc)
        

    elif batch_info == 'ME' and shift_info == '2':
        loc = './excel/ME2.xlsx'
        xfile = openpyxl.load_workbook(loc)
        sheet = xfile.get_sheet_by_name('Sheet1')    
        row=10
        col=1
        for data in batch_data:
            Main.append(data)
        print (Main)
        for entry in Main:
            print(entry)
            sheet.cell(row, col).value=entry
            col=col+1
            if col == 8:
                col = 1
                row = row + 1
                if row == 14:
                    row = row+1
        xfile.save(loc)
        
    return loc

def excel_location(location_info):
    """ Preparing excel sheet for location wise timetable
        
        Args:
        1. location_info - Data from master timetable of particular location i.e. classroom number
    """
    Main=[]
    loc = './excel/Location.xlsx'
    xfile = openpyxl.load_workbook(loc)
    sheet = xfile.get_sheet_by_name('Sheet1')    
    row=10
    col=1
    for data in location_info:
        Main.append(data)
    print (Main)
    for entry in Main:
        print(entry)
        sheet.cell(row, col).value=entry
        col=col+1
        if col == 8:
            col = 1
            row = row + 1
            if row == 14:
                row = row+1

    xfile.save(loc)
    return loc

def excel_teacher(teacher_info):
    """ Preparing excel sheet for faculty wise timetable
        
        Args:
        1. teacher_info - Data from master timetable of particular teacher 
    """
    Main=[]
    loc = './excel/Teacher.xlsx'
    xfile = openpyxl.load_workbook(loc)
    sheet = xfile.get_sheet_by_name('Sheet1')    
    row=10
    col=1
    for data in teacher_info:
        Main.append(data)
    print (Main)
    for entry in Main:
        print(entry)
        sheet.cell(row, col).value=entry
        col=col+1
        if col == 8:
            col = 1
            row = row + 1
            if row == 14:
                row = row+1

    xfile.save(loc)
    return loc

def excel_master(master_info):
    """ Preparing excel sheet for master timetable
        
        Args:
        1. master_info - Data of master timetable 
    """
    Main=[]
    loc = './excel/Master.xlsx'
    xfile = openpyxl.load_workbook(loc)
    sheet = xfile.get_sheet_by_name('Sheet1')    
    row=9
    col=2
    for data in master_info:
        Main.append(data)
    print (Main)
    for entry in Main:
        print(entry)
        sheet.cell(row, col).value=entry
        col=col+1
        if col == 18:
            col = 2
            row = row + 1
    xfile.save(loc)
    return loc
