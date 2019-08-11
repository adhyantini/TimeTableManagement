import openpyxl
from openpyxl import load_workbook

def upload_excel():
    sheet_name = 'Sheet1'
    wb = openpyxl.load_workbook(my_excel)
    sheet = wb.get_sheet_by_name(sheet_name)
    col = 1
    row = 2
    excel_contents=[]
    for i in range(row,sheet.max_row+1):
        Day = sheet.cell(row,col).value
        Slot = sheet.cell(row,col+1).value
        String = sheet.cell(row,col+2).value
        Batch = String.split('-',3)
        print (Batch)
